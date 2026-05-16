"""
excel_manager.py - Module xử lý dữ liệu từ Excel Local (.xlsx) hoặc Google Sheets Online.

Hỗ trợ:
    1. Đọc dữ liệu từ file local hoặc link Google Sheet.
    2. Kiểm tra tính hợp lệ của media (file/thư mục).
    3. Cập nhật Status (DONE) trực tiếp lên Cloud hoặc file Local.
"""

import os
import pandas as pd
from openpyxl import load_workbook
from colorama import Fore, Style
import gspread
from oauth2client.service_account import ServiceAccountCredentials

# Các định dạng file ảnh/video được hỗ trợ
DINH_DANG_ANH = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp'}
DINH_DANG_VIDEO = {'.mp4', '.mov', '.avi', '.mkv', '.webm', '.flv'}
DINH_DANG_MEDIA = DINH_DANG_ANH | DINH_DANG_VIDEO

# Danh sách các cột BẮT BUỘC
REQUIRED_COLUMNS = ['Ma_Bai_Dang', 'Link_Bai_Dang', 'Caption', 'Anh_Video', 'Status']

def ket_noi_google_sheets():
    """Thiết lập kết nối với Google Sheets API bằng credentials.json"""
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    try:
        creds = ServiceAccountCredentials.from_json_keyfile_name("credentials.json", scope)
        client = gspread.authorize(creds)
        return client
    except Exception as e:
        print(f"{Fore.RED}[LỖI] Không thể kết nối Google Sheets: {e}{Style.RESET_ALL}")
        return None

def doc_du_lieu(nguon: str) -> pd.DataFrame | None:
    """
    Đọc dữ liệu từ file Local (.xlsx) hoặc Google Sheets ID.
    
    Args:
        nguon: Đường dẫn file local (data/...) hoặc ID của Google Sheet.
    """
    # KIỂM TRA NẾU LÀ GOOGLE SHEETS
    is_google_sheet = "docs.google.com" in nguon or (nguon.startswith("http") and "spreadsheets" in nguon)

    if is_google_sheet:
        print(f"{Fore.CYAN}🌐 Đang kết nối tới Google Sheets Online...{Style.RESET_ALL}")
        client = ket_noi_google_sheets()
        if not client: return None
        try:
            # Nếu là link đầy đủ, trích xuất ID
            sheet_id = nguon
            if "docs.google.com" in nguon:
                sheet_id = nguon.split("/d/")[1].split("/")[0]
            
            sh = client.open_by_key(sheet_id)
            worksheet = sh.get_worksheet(0) # Lấy sheet đầu tiên
            data = worksheet.get_all_records()
            df = pd.DataFrame(data)
            
            # Đảm bảo các cột trống không bị bỏ qua
            for col in REQUIRED_COLUMNS:
                if col not in df.columns:
                    df[col] = ""
        except Exception as e:
            import traceback
            print(f"{Fore.RED}[LỖI] Lỗi khi truy cập Google Sheet:{Style.RESET_ALL}")
            print(f"{Fore.RED}  Loại lỗi: {type(e).__name__}{Style.RESET_ALL}")
            print(f"{Fore.RED}  Chi tiết: {e}{Style.RESET_ALL}")
            traceback.print_exc()
            return None
    else:
        # ĐỌC FILE LOCAL (.xlsx)
        if not os.path.exists(nguon):
            print(f"{Fore.RED}[LỖI] Không tìm thấy file local: {nguon}{Style.RESET_ALL}")
            return None
        try:
            df = pd.read_excel(nguon, engine='openpyxl', dtype=str)
        except Exception as e:
            print(f"{Fore.RED}[LỖI] Lỗi khi đọc file Excel local: {e}{Style.RESET_ALL}")
            return None

    # --- Kiểm tra cột bắt buộc ---
    cot_thieu = [cot for cot in REQUIRED_COLUMNS if cot not in df.columns]
    if cot_thieu:
        print(f"{Fore.RED}[LỖI] Thiếu các cột: {', '.join(cot_thieu)}{Style.RESET_ALL}")
        return None

    # --- Lọc bài chưa làm (Status != DONE) ---
    df['Status'] = df['Status'].fillna('').astype(str).str.strip().str.upper()
    df_chua_lam = df[df['Status'] != 'DONE'].copy()
    
    if len(df_chua_lam) == 0:
        print(f"{Fore.YELLOW}[THÔNG BÁO] Tất cả bài đăng đều đã hoàn thành (DONE).{Style.RESET_ALL}")
        return None

    df_chua_lam = df_chua_lam.reset_index(drop=True)
    print(f"{Fore.GREEN}✅ Đọc thành công: {len(df_chua_lam)} bài đăng cần xử lý.{Style.RESET_ALL}")
    return df_chua_lam

def tach_danh_sach(chuoi_gop: str) -> list[str]:
    """Tách chuỗi bằng dấu | hoặc xuống dòng (\\n). Hỗ trợ cả 2 kiểu nhập."""
    if pd.isna(chuoi_gop) or str(chuoi_gop).strip() == '' or str(chuoi_gop).strip().lower() == 'nan':
        return []
    import re
    return [item.strip() for item in re.split(r'[|\n]', str(chuoi_gop)) if item.strip()]

def phan_giai_media(chuoi_media: str) -> list[str]:
    danh_sach_tho = tach_danh_sach(chuoi_media)
    if not danh_sach_tho: return []
    ket_qua = []
    for duong_dan in danh_sach_tho:
        if os.path.isdir(duong_dan):
            for ten_file in sorted(os.listdir(duong_dan)):
                if os.path.splitext(ten_file)[1].lower() in DINH_DANG_MEDIA:
                    ket_qua.append(os.path.join(duong_dan, ten_file))
        else:
            ket_qua.append(duong_dan)
    return ket_qua

def kiem_tra_file_media(danh_sach_media: list[str], ma_bai: str) -> tuple[bool, list[str]]:
    file_khong_ton_tai = [f for f in danh_sach_media if not os.path.exists(f)]
    if file_khong_ton_tai:
        print(f"{Fore.RED}[LỖI MEDIA - {ma_bai}] Không tìm thấy:{Style.RESET_ALL}")
        for f in file_khong_ton_tai: print(f"{Fore.RED}  ✗ {f}{Style.RESET_ALL}")
        return False, file_khong_ton_tai
    return True, []

def cap_nhat_status(nguon: str, ma_bai: str, gia_tri: str):
    """Cập nhật Status vào Google Sheets hoặc File Local dựa trên Mã Bài Đăng"""
    is_google_sheet = "docs.google.com" in nguon or (nguon.startswith("http") and "spreadsheets" in nguon)

    if is_google_sheet:
        client = ket_noi_google_sheets()
        if not client: return
        try:
            sheet_id = nguon
            if "docs.google.com" in nguon:
                sheet_id = nguon.split("/d/")[1].split("/")[0]
            sh = client.open_by_key(sheet_id)
            worksheet = sh.get_worksheet(0)
            
            # Tìm dòng có mã bài tương ứng
            cell = worksheet.find(ma_bai)
            if cell:
                # Tìm cột Status
                header = worksheet.row_values(1)
                try:
                    status_col = header.index("Status") + 1
                    worksheet.update_cell(cell.row, status_col, gia_tri)
                except ValueError:
                    print(f"{Fore.RED}Không tìm thấy cột Status trên Google Sheets!{Style.RESET_ALL}")
        except Exception as e:
            print(f"{Fore.YELLOW}[CẢNH BÁO] Không thể cập nhật Status lên Google Sheets: {e}{Style.RESET_ALL}")
    else:
        # Cập nhật file Local .xlsx
        try:
            wb = load_workbook(nguon)
            ws = wb.active
            # Tìm vị trí cột Status và dòng Ma_Bai_Dang
            status_col = None
            for col_idx, cell in enumerate(ws[1], start=1):
                if str(cell.value).strip() == 'Status':
                    status_col = col_idx
                    break
            
            if status_col:
                for row_idx in range(2, ws.max_row + 1):
                    if str(ws.cell(row=row_idx, column=1).value).strip() == ma_bai:
                        ws.cell(row=row_idx, column=status_col, value=gia_tri)
                        break
            wb.save(nguon)
            wb.close()
        except Exception as e:
            print(f"{Fore.YELLOW}[CẢNH BÁO] Lỗi cập nhật file local: {e}{Style.RESET_ALL}")

def tien_kiem_tra_toan_bo(df: pd.DataFrame) -> tuple[bool, list[str]]:
    print(f"\n{Fore.CYAN}🔍 Đang kiểm tra tính hợp lệ dữ liệu...{Style.RESET_ALL}")
    bai_loi = []
    for idx, row in df.iterrows():
        ma_bai = str(row.get('Ma_Bai_Dang', '')).strip()
        if not ma_bai or ma_bai.lower() == 'nan':
            bai_loi.append(f"Dòng {idx+1}")
            continue
        links = tach_danh_sach(row.get('Link_Bai_Dang', ''))
        if not links:
            bai_loi.append(ma_bai)
            continue
        media = phan_giai_media(row.get('Anh_Video', ''))
        if media:
            ok, _ = kiem_tra_file_media(media, ma_bai)
            if not ok: bai_loi.append(ma_bai)
    
    if bai_loi:
        print(f"{Fore.RED}❌ PHÁT HIỆN LỖI ở: {', '.join(bai_loi)}{Style.RESET_ALL}")
        return False, bai_loi
    print(f"{Fore.GREEN}✅ Dữ liệu hợp lệ.{Style.RESET_ALL}")
    return True, []
